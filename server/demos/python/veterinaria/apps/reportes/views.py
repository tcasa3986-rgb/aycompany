from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum
from apps.clientes.models import Cliente, Mascota
from apps.facturacion.models import Factura, DetalleFactura
from apps.inventario.models import Producto
from apps.medico.models import Vacuna
import io

# ReportLab para PDF
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# openpyxl para Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
def reportes_index(request):
    return render(request, 'reportes/index.html')


@login_required
def reporte_clientes_pdf(request):
    clientes = Cliente.objects.prefetch_related('mascotas').order_by('apellidos', 'nombres')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Título
    titulo_style = ParagraphStyle('Titulo', parent=styles['Title'], fontSize=16, spaceAfter=12)
    story.append(Paragraph("Reporte de Clientes y Mascotas", titulo_style))
    story.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    # Tabla
    data = [['N°', 'Nombre', 'DNI', 'Teléfono', 'Email', 'Mascotas', 'Registro']]
    for i, c in enumerate(clientes, 1):
        data.append([
            str(i),
            c.nombre_completo,
            c.dni,
            c.telefono,
            c.email or '-',
            str(c.mascotas.count()),
            c.fecha_registro.strftime('%d/%m/%Y'),
        ])

    tabla = Table(data, colWidths=[1*cm, 5*cm, 2.5*cm, 3*cm, 4*cm, 2*cm, 2.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0bb8c8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fafe')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0e8ec')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(tabla)
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_clientes.pdf"'
    return response


@login_required
def reporte_ventas_excel(request):
    facturas = Factura.objects.filter(estado='PAGADA').select_related('cliente').order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas'

    header_fill = PatternFill(start_color='0BB8C8', end_color='0BB8C8', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    alt_fill = PatternFill(start_color='EEF9FB', end_color='EEF9FB', fill_type='solid')

    headers = ['N° Factura', 'Fecha', 'Cliente', 'Método Pago', 'Subtotal', 'IGV', 'Total', 'Estado']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, f in enumerate(facturas, 2):
        fill = alt_fill if row % 2 == 0 else PatternFill()
        values = [
            f.numero,
            f.fecha.strftime('%d/%m/%Y %H:%M'),
            f.cliente.nombre_completo if f.cliente else 'Anónimo',
            f.get_metodo_pago_display(),
            float(f.subtotal),
            float(f.igv),
            float(f.total),
            f.get_estado_display(),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    for col in ['E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14

    # Totales
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=5, value=sum(float(f.subtotal) for f in facturas)).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=sum(float(f.total) for f in facturas)).font = Font(bold=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    wb.save(response)
    return response


@login_required
def reporte_inventario_pdf(request):
    productos = Producto.objects.filter(is_active=True).select_related('categoria', 'proveedor').order_by('nombre')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Inventario", ParagraphStyle('T', parent=styles['Title'], fontSize=16)))
    story.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    data = [['Código', 'Producto', 'Categoría', 'Stock', 'Stock Mín.', 'P. Compra', 'P. Venta', 'Estado']]
    for p in productos:
        estado = '⚠ Bajo' if p.necesita_reposicion else 'OK'
        data.append([
            p.codigo or '-',
            p.nombre[:35],
            p.categoria.nombre if p.categoria else '-',
            str(p.stock_actual),
            str(p.stock_minimo),
            f"S/ {p.precio_compra}",
            f"S/ {p.precio_venta}",
            estado,
        ])

    tabla = Table(data, colWidths=[2*cm, 5*cm, 3*cm, 1.5*cm, 1.8*cm, 2*cm, 2*cm, 1.5*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fffbeb')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tabla)
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'
    return response


@login_required
def reporte_vacunas_pdf(request):
    vacunas = Vacuna.objects.select_related('mascota__cliente', 'veterinario').order_by('-fecha_aplicacion')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Reporte de Vacunaciones", ParagraphStyle('T', parent=styles['Title'], fontSize=16)))
    story.append(Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 0.5*cm))

    data = [['Mascota', 'Cliente', 'Vacuna', 'Fecha Aplic.', 'Próxima Dosis', 'Veterinario']]
    for v in vacunas:
        data.append([
            v.mascota.nombre,
            v.mascota.cliente.nombre_completo,
            v.nombre_vacuna,
            v.fecha_aplicacion.strftime('%d/%m/%Y'),
            v.fecha_proxima_dosis.strftime('%d/%m/%Y') if v.fecha_proxima_dosis else '-',
            v.veterinario.get_full_name() if v.veterinario else '-',
        ])

    tabla = Table(data, colWidths=[3*cm, 4.5*cm, 4*cm, 2.5*cm, 2.5*cm, 3*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f3ff')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tabla)
    doc.build(story)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_vacunas.pdf"'
    return response
